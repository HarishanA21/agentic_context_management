import base64
import io
import os
from pathlib import Path

from langchain.tools import tool
from langchain_core.runnables import RunnableConfig

from Tools._paths import get_session_ids, safe_name
from storage import file_key, get_bucket, is_not_found

MAX_READ_BYTES = 200_000  # 200 KB cap to keep tool responses small

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}


def _read_pdf(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "Error: pypdf not installed (pip install pypdf)."
    reader = PdfReader(io.BytesIO(data))
    out = []
    for i, page in enumerate(reader.pages, 1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        out.append(f"--- Page {i} ---\n{text}")
    return "\n\n".join(out)


def _read_docx(data: bytes) -> str:
    try:
        from docx import Document
    except ImportError:
        return "Error: python-docx not installed (pip install python-docx)."
    doc = Document(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _read_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return "Error: openpyxl not installed (pip install openpyxl)."
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        out.append(f"--- Sheet: {sheet_name} ---")
        for row in ws.iter_rows(values_only=True):
            out.append("\t".join("" if c is None else str(c) for c in row))
    return "\n".join(out)


def _read_image(data: bytes, ext: str) -> str:
    """Describe an image and extract any visible text using a vision LLM.

    Calls a multimodal model on OpenRouter so the agent can answer questions
    about images even though our main chat model isn't multimodal.
    """
    api_key = os.getenv("OPENROUTER_API_KEY")
    if not api_key:
        return "Error: OPENROUTER_API_KEY not set; cannot analyze images."
    try:
        from langchain_core.messages import HumanMessage
        from langchain_openai import ChatOpenAI
    except ImportError:
        return "Error: langchain not installed."

    mime_map = {
        ".jpg": "jpeg",
        ".jpeg": "jpeg",
        ".png": "png",
        ".webp": "webp",
        ".gif": "gif",
    }
    mime = mime_map.get(ext, "jpeg")
    b64 = base64.b64encode(data).decode("ascii")
    data_url = f"data:image/{mime};base64,{b64}"

    model_name = os.getenv(
        "VISION_MODEL", "meta-llama/llama-3.2-11b-vision-instruct:free"
    )
    try:
        client = ChatOpenAI(
            model=model_name,
            openai_api_key=api_key,
            openai_api_base="https://openrouter.ai/api/v1",
            max_tokens=1500,
            temperature=0.2,
            default_headers={
                "HTTP-Referer": "http://localhost",
                "X-Title": "FYP Agent — vision",
            },
        )
        msg = HumanMessage(
            content=[
                {
                    "type": "text",
                    "text": (
                        "Describe this image in detail. If it contains any "
                        "text (signs, slides, handwriting, screenshots), "
                        "transcribe the text exactly. Be thorough but concise."
                    ),
                },
                {"type": "image_url", "image_url": {"url": data_url}},
            ]
        )
        result = client.invoke([msg])
        return f"[Image: {ext.lstrip('.').upper()}]\n\n{result.content}"
    except Exception as e:
        emsg = str(e)
        if "429" in emsg or "rate" in emsg.lower():
            return "Error: vision model rate-limited. Try again shortly."
        return f"Error analyzing image: {emsg[:200]}"


@tool
def read_project_file(filename: str, config: RunnableConfig) -> str:
    """Read a file from the current project's uploaded files.

    Supports plain text (UTF-8), PDF (.pdf), Word (.docx), Excel (.xlsx),
    and images (.jpg, .jpeg, .png, .webp, .gif). For images, returns a
    description of the image content plus any visible text.
    Long files are truncated to fit within the response budget.

    Args:
        filename: The basename of the file to read (no directory components).
    """
    try:
        user_id, session_id = get_session_ids(config)
        name = safe_name(filename)
    except ValueError as e:
        return f"Error: {e}"

    try:
        data: bytes = get_bucket().download(file_key(user_id, session_id, name))
    except Exception as e:
        if is_not_found(e):
            return f"Error: '{name}' not found in project files."
        return f"Error: failed to download file: {e}"

    ext = Path(name).suffix.lower()
    try:
        if ext == ".pdf":
            text = _read_pdf(data)
        elif ext == ".docx":
            text = _read_docx(data)
        elif ext == ".xlsx":
            text = _read_xlsx(data)
        elif ext in IMAGE_EXTS:
            text = _read_image(data, ext)
        else:
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                return (
                    f"Error: '{name}' is not a UTF-8 text file. "
                    "Supported binary formats: .pdf, .docx, .xlsx, "
                    ".jpg, .jpeg, .png, .webp, .gif."
                )
    except Exception as e:
        return f"Error reading {ext or 'file'}: {e}"

    if len(text.encode("utf-8")) > MAX_READ_BYTES:
        text = text[: MAX_READ_BYTES // 2] + "\n\n[... truncated ...]"
    return text
